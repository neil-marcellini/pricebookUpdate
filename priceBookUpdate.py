# Neil Marcellini
# December 2019

import os
import math
import time
from openpyxl import load_workbook

multiplier = 1.05


def roundUpNearestFifty(value):
    rval = value
    val_string = str(value)
    dot_index = val_string.find(".")
    # if not an integer
    if dot_index > 0:
        # look at digit to right
        tens_digit = int(val_string[dot_index + 1])
        if tens_digit >= 5:
            # round up to next integer
            rval = math.ceil(value)
        else:
            # round up to .5
            rval = math.ceil(value) / 2 + int(value) / 2
    return rval


def getFileNames():
    # a list of all the excel files in the pricebook directory
    file_names = []
    for root, dirs, files in os.walk("."):
        for file_ in files:
            if ".xlsx" in file_ and "~$" not in file_ and "Cover" not in file_:
                file_names.append(os.path.join(root, file_))
    return file_names

def updateWorkbook(name):
    wb = load_workbook(name)
    ws = wb['wholesale']
    for row in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if type(cell.value) == int or type(cell.value) == float:
                new_val = cell.value * multiplier
                new_val = roundUpNearestFifty(new_val)
                cell.value = new_val
    #new_name = name.rstrip(".xlsx") + "_updated" + ".xlsx"
    wb.save(name)

def main():
    start_time = time.time()
    file_names = getFileNames()
    for name in file_names:
        updateWorkbook(name)
    print("Total time: %s" % (time.time() - start_time))

main()
